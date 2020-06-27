#opening the data sheet
import openpyxl
import math

#for each location on the sheet we have to calculate distance to all other
#locations, these distances must be written to sheets in the workbook
#beginning processing
wb = openpyxl.load_workbook('C:/Users/neelu/OneDrive/Desktop/Computer Science/Pune Food Tour/place_data.xlsx')
main_sheet = wb.get_sheet_by_name('Sheet0') #original sheet exported from google maps
starting_location = main_sheet.cell(row = 1, column = 1).value
#quick function for distance between two points
def calculateDistance(x1,y1,x2,y2):
     dist = math.sqrt((x2 - x1)**2 + (y2 - y1)**2)
     return dist

def pre_processing(wb, main_sheet):
    for i in range(1 , 25): #len(wb['A'])+1
        first_x_location = 'B' + str(i)
        first_x = main_sheet[first_x_location].value
        first_y_location = 'C' + str(i)
        first_y = main_sheet[first_y_location].value
        #create sheet for the distances between all other destinations and i'th one
        name_of_new_sheet = 'Sheet' + str(i)
        wb.create_sheet(index = i, title = name_of_new_sheet)
        current_sheet = wb.get_sheet_by_name(name_of_new_sheet)

        for j in range(1 , 25): #len(wb['A'])+1
            second_x_location = 'B' + str(j)
            second_x = main_sheet[second_x_location].value
            second_y_location = 'C' + str(j)
            second_y = main_sheet[second_y_location].value
            temp = calculateDistance(float(first_x),float(first_y),float(second_x),float(second_y))
            current_cell = 'A' + str(j)
            from_cell = 'A' + str(i)
            #writing something like 'Home to Yewale'
            current_sheet.cell(row = j, column = 1).value = main_sheet[from_cell].value + ' to ' + main_sheet[current_cell].value
            #writing the distance
            current_cell = 'B' + str(j)
            current_sheet.cell(row = j, column = 2).value = temp
    wb.save('C:/Users/neelu/OneDrive/Desktop/Computer Science/Pune Food Tour/place_data2.xlsx')

def greedy_algorithm(wb, main_sheet):
    route = []
    indices = list(range(2, 25))
    route.append(starting_location)
    count = 1 #number of iterations we've made
    i = 1; #i is the sheet number we're looking at
    while count < 24:
        name_of_new_sheet = 'Sheet' + str(i)
        current_sheet = wb.get_sheet_by_name(name_of_new_sheet)
        temp = current_sheet.cell(row = indices[0], column = 2).value
        for j in range (1, 25):
            distance = current_sheet.cell(row = j, column = 2).value
            if distance <= temp and distance!=0 and main_sheet.cell(row = j, column = 1).value not in route:
                temp = distance
                index = j
        route.append(main_sheet.cell(row = index, column = 1).value)
        i = index;
        count = count + 1;
        if index in indices:
            indices.remove(index)
    route.append(starting_location)
    print(route)

#pre_processing(wb, main_sheet)
greedy_algorithm(wb, main_sheet)
